from pathlib import Path
from openpyxl import load_workbook
import json
from datetime import datetime, date

src = Path("data/mlit_r5_samples_unverified/02_tohoku_kouji_2304.xlsx")
out = Path("data/mlit_r5_confirmed/tohoku_r5_kouji_contracts.json")

wb = load_workbook(src, data_only=True)
ws = wb[wb.sheetnames[0]]

def conv(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return v

records = []

for row in ws.iter_rows(min_row=7, values_only=True):
    if not row or not row[1]:
        continue

    records.append({
        "bureau": "東北地方整備局",
        "fiscal_year": "R5",
        "month": "2023-04",
        "category": "工事",
        "source_type": "excel",
        "department": conv(row[0]),
        "project_name": conv(row[1]),
        "project_location": conv(row[2]),
        "bid_date": conv(row[3]),
        "contract_date": conv(row[4]),
        "contract_period": conv(row[5]),
        "work_type": conv(row[6]),
        "bid_method": conv(row[7]),
        "comprehensive_evaluation": conv(row[8]),
        "bidder": conv(row[9]),
        "bidder_address": conv(row[10]),
        "planned_price_ex_tax": conv(row[11]),
        "investigation_base_price_ex_tax": conv(row[12]),
        "score": conv(row[13]),
        "bid_amount_1st_ex_tax": conv(row[14]),
        "evaluation_value_1st": conv(row[15]),
        "bid_amount_2nd_ex_tax": conv(row[16]),
        "evaluation_value_2nd": conv(row[17]),
        "bid_amount_3rd_ex_tax": conv(row[18]),
        "evaluation_value_3rd": conv(row[19]),
        "estimate_amount_ex_tax": conv(row[20]),
        "winning_rate": conv(row[21]),
        "remarks": conv(row[22]),
        "is_awarded": conv(row[22]) == "落札",
    })

out.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
print("saved:", out)
print("count:", len(records))
print("awarded:", sum(1 for r in records if r["is_awarded"]))
