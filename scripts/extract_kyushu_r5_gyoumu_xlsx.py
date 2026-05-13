from pathlib import Path
from openpyxl import load_workbook
import json
from datetime import datetime, date

src = Path("data/mlit_r5_confirmed/kyushu_r5_gyoumu_2304.xlsx")
out = Path("data/mlit_r5_confirmed/kyushu_r5_gyoumu_contracts.json")

wb = load_workbook(src, data_only=True)
ws = wb["業務（４月）"]

def conv(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if v == "":
        return None
    return v

records = []

for row in ws.iter_rows(min_row=4, values_only=True):
    if not row or not row[1]:
        continue

    remarks = conv(row[21])
    if remarks not in ["落札", "決定"]:
        continue

    records.append({
        "bureau": "九州地方整備局",
        "fiscal_year": "R5",
        "month": "2023-04",
        "category": "業務",
        "source_type": "excel",
        "department": conv(row[0]),
        "project_name": conv(row[1]),
        "bid_date": conv(row[2]),
        "contract_date": conv(row[3]),
        "work_type": conv(row[4]),
        "bid_method": conv(row[5]),
        "comprehensive_evaluation": conv(row[6]),
        "bidder": conv(row[7]),
        "planned_price_ex_tax": conv(row[8]),
        "investigation_base_price_ex_tax": conv(row[9]),
        "technical_score": conv(row[10]),
        "bid_amount_1st_ex_tax": conv(row[11]),
        "price_score_1st": conv(row[12]),
        "evaluation_value_1st": conv(row[13]),
        "bid_amount_2nd_ex_tax": conv(row[14]),
        "price_score_2nd": conv(row[15]),
        "evaluation_value_2nd": conv(row[16]),
        "bid_amount_3rd_ex_tax": conv(row[17]),
        "price_score_3rd": conv(row[18]),
        "evaluation_value_3rd": conv(row[19]),
        "estimate_amount_ex_tax": conv(row[20]),
        "remarks": remarks,
        "is_awarded": True,
    })

out.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
print("saved:", out)
print("count:", len(records))
