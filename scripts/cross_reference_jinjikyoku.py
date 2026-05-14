#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 内閣人事局の再就職データ x デジタル庁R7受注企業 クロスリファレンス
# データ: 内閣人事局 国家公務員法106条の25第2項 令和5年度公表分

import urllib.request
import urllib.error
from pathlib import Path
import openpyxl
import json
import sys

BASE = Path("/Volumes/SN0W8ALL/tokubetsu-kaikei")
DATA_DIR = BASE / "data" / "jinjikyoku_r5"
OUTPUT_DIR = BASE / "output"
DATA_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

FILES = {
    "在職中の届出": "https://www.cas.go.jp/jp/gaiyou/jimu/jinjikyoku/106-25-2/r06/files/siryou1-1.xlsx",
    "離職前の事前届出": "https://www.cas.go.jp/jp/gaiyou/jimu/jinjikyoku/106-25-2/r06/files/siryou2.xlsx",
    "離職後の事後届出": "https://www.cas.go.jp/jp/gaiyou/jimu/jinjikyoku/106-25-2/r06/files/siryou3-1.xlsx",
}

print("=" * 70)
print("内閣人事局 x デジタル庁R7受注企業 クロスリファレンス")
print("データ: 令和5年4月1日 - 令和6年3月31日の再就職状況（計1,742件）")
print("出典: 内閣人事局 国家公務員法106条の25第2項")
print("=" * 70)

# === [1/5] ダウンロード ===
print("")
print("[1/5] Excel ダウンロード")
for name, url in FILES.items():
    fname = url.split("/")[-1]
    path = DATA_DIR / fname
    if path.exists():
        print("  (cached) {}: {:,} bytes".format(fname, path.stat().st_size))
        continue
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as r:
            path.write_bytes(r.read())
        print("  OK {}: {:,} bytes".format(fname, path.stat().st_size))
    except Exception as e:
        print("  FAIL {}: {}".format(fname, e))
        sys.exit(1)

# === デジタル庁R7受注企業18社（精密マッチング用キーワード） ===
DIGITAL_R7_VENDORS = {
    "日本電気（NEC）": ["日本電気株式会社"],
    "NTTデータ": ["株式会社ＮＴＴデータ", "株式会社エヌ・ティ・ティ・データ", "ＮＴＴデータ株式会社", "NTTデータ株式会社"],
    "NTT東日本（東日本電信電話）": ["東日本電信電話株式会社"],
    "KDDI": ["KDDI株式会社", "ＫＤＤＩ株式会社"],
    "日本マイクロソフト": ["日本マイクロソフト株式会社"],
    "NECフィールディング": ["ＮＥＣフィールディング株式会社", "NECフィールディング株式会社"],
    "アクセンチュア": ["アクセンチュア株式会社"],
    "富士ソフト": ["富士ソフト株式会社"],
    "富士通": ["富士通株式会社"],
    "行政情報システム研究所": ["一般社団法人行政情報システム研究所"],
    "PwCコンサルティング": ["PwCコンサルティング合同会社", "ＰｗＣコンサルティング合同会社"],
    "フューチャーアーキテクト": ["フューチャーアーキテクト株式会社"],
    "日立製作所": ["株式会社日立製作所"],
    "富士通ネットワーク": ["富士通ネットワークソリューションズ"],
    "NTTコミュニケーションズ／NTTドコモビジネス": [
        "エヌ・ティ・ティ・コミュニケーションズ",
        "ＮＴＴコミュニケーションズ",
        "NTTコミュニケーションズ",
        "ＮＴＴドコモビジネス",
        "NTTドコモビジネス",
    ],
    "SBテクノロジー": ["SBテクノロジー株式会社", "ＳＢテクノロジー株式会社"],
    "シンプレクス": ["シンプレクス株式会社"],
    "デロイトトーマツ": ["デロイトトーマツコンサルティング"],
}

# === [2/5] Excel パース ===
print("")
print("[2/5] Excel パース")


def parse_excel(path, source_label):
    wb = openpyxl.load_workbook(path, data_only=True)
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        # ヘッダー行を発見
        header_idx = None
        for i, row in enumerate(rows[:15]):
            if not row:
                continue
            text = " ".join(str(c) for c in row if c is not None)
            if "氏名" in text or "再就職" in text or "法人" in text or "省庁" in text:
                header_idx = i
                break
        if header_idx is None:
            header_idx = 0
        headers = [
            str(c).strip() if c is not None else "col_{}".format(j)
            for j, c in enumerate(rows[header_idx])
        ]
        for row in rows[header_idx + 1:]:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            record = {}
            for j, val in enumerate(row):
                if j < len(headers):
                    key = headers[j] if headers[j] else "col_{}".format(j)
                    record[key] = str(val).strip() if val is not None else ""
            record["__sheet"] = sheet_name
            record["__source"] = source_label
            records.append(record)
    return records


all_records = []
for label, url in FILES.items():
    fname = url.split("/")[-1]
    path = DATA_DIR / fname
    recs = parse_excel(path, label)
    print("  {}: {} 件".format(label, len(recs)))
    all_records.extend(recs)

print("  合計: {} 件".format(len(all_records)))

# === [3/5] JOIN 実行 ===
print("")
print("[3/5] JOIN 実行 (18社の精密マッチング)")


def match_vendor(record):
    blob = " ".join(str(v) for k, v in record.items() if not k.startswith("__"))
    matched = []
    for vendor, keywords in DIGITAL_R7_VENDORS.items():
        for kw in keywords:
            if kw in blob:
                matched.append(vendor)
                break
    return matched


matches = []
for record in all_records:
    vendors = match_vendor(record)
    if vendors:
        matches.append({"matched_vendors": vendors, "record": record})

print("  マッチ件数: {} 件".format(len(matches)))

# === [4/5] 集計 ===
vendor_counts = {}
for m in matches:
    for v in m["matched_vendors"]:
        vendor_counts[v] = vendor_counts.get(v, 0) + 1

print("")
print("[4/5] ベンダー別 集計")
print("  " + "-" * 60)
print("  {:<40} | {:>5}".format("受注企業（デジタル庁R7）", "件数"))
print("  " + "-" * 60)
for vendor, count in sorted(vendor_counts.items(), key=lambda x: -x[1]):
    print("  {:<40} | {:>5}".format(vendor, count))
print("  " + "-" * 60)

# === [5/5] 出力 ===
print("")
print("[5/5] 出力")

output = {
    "source": "内閣人事局 国家公務員法106条の25第2項に基づく再就職状況の公表（令和5年度）",
    "source_url": "https://www.cas.go.jp/jp/gaiyou/jimu/jinjikyoku/106-25-2/r06/kouhyou_0924.html",
    "period": "令和5年4月1日 - 令和6年3月31日",
    "total_records": len(all_records),
    "match_count": len(matches),
    "vendor_summary": vendor_counts,
    "matches": matches,
}

out_path = OUTPUT_DIR / "digital_r7_jinjikyoku_cross_r5.json"
out_path.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
print("  JSON: {} ({:,} bytes)".format(out_path.relative_to(BASE), out_path.stat().st_size))

# HTML レポート
lines = []
lines.append("<!DOCTYPE html><html lang='ja'><head><meta charset='utf-8'>")
lines.append("<title>内閣人事局 x デジタル庁R7 クロスリファレンス</title>")
lines.append("<style>")
lines.append("body{font-family:'Hiragino Sans',sans-serif;background:#0d1117;color:#e6edf3;padding:24px;max-width:1200px;margin:auto;line-height:1.7}")
lines.append("h1{color:#58a6ff;border-bottom:1px solid #30363d;padding-bottom:8px}")
lines.append("h2{color:#7ee787;margin-top:30px}")
lines.append("table{border-collapse:collapse;margin:14px 0;width:100%}")
lines.append("th,td{border:1px solid #30363d;padding:8px 12px;text-align:left}")
lines.append("th{background:#161b22;color:#79c0ff}")
lines.append(".match{background:#161b22;padding:14px;margin:10px 0;border-left:4px solid #f97316;font-size:13px;border-radius:4px}")
lines.append(".match b{color:#ffa657}")
lines.append("a{color:#58a6ff}")
lines.append(".count-big{font-size:32px;color:#ff7b72;font-weight:bold}")
lines.append(".meta{color:#8b949e;font-size:12px}")
lines.append("</style></head><body>")
lines.append("<h1>内閣人事局 x デジタル庁R7受注企業 クロスリファレンス</h1>")
lines.append("<p class='meta'>データソース: <a href='https://www.cas.go.jp/jp/gaiyou/jimu/jinjikyoku/106-25-2/r06/kouhyou_0924.html' target='_blank'>内閣人事局 令和5年度公表分</a>")
lines.append("（{} 件、 令和5年4月1日-令和6年3月31日）</p>".format(len(all_records)))
lines.append("<div>マッチ件数: <span class='count-big'>{}</span> 件</div>".format(len(matches)))

lines.append("<h2>ベンダー別 集計</h2>")
lines.append("<table><tr><th>受注企業（デジタル庁R7）</th><th>再就職件数</th></tr>")
for vendor, count in sorted(vendor_counts.items(), key=lambda x: -x[1]):
    lines.append("<tr><td>{}</td><td>{}</td></tr>".format(vendor, count))
lines.append("</table>")

if not vendor_counts:
    lines.append("<p style='color:#f97316'>※ 該当ベンダーへの再就職はR5公表分には記録されていません。 マッチング条件を緩める、 R7四半期データを追加する、 等の検討を。</p>")

lines.append("<h2>詳細マッチ（最大50件）</h2>")
for i, m in enumerate(matches[:50]):
    lines.append("<div class='match'>")
    lines.append("<b>#{} 受注企業: {}</b><br>".format(i + 1, " / ".join(m["matched_vendors"])))
    lines.append("<small class='meta'>ソース: {} / シート: {}</small><br>".format(
        m["record"].get("__source", ""), m["record"].get("__sheet", "")
    ))
    for k, v in m["record"].items():
        if k.startswith("__") or not v:
            continue
        lines.append("<small>{}: {}</small><br>".format(k, str(v)[:300]))
    lines.append("</div>")

if len(matches) > 50:
    lines.append("<p class='meta'>※ 残り {} 件は JSON ファイルを参照</p>".format(len(matches) - 50))

lines.append("<hr><p class='meta'>※ 本レポートは公開データの相関を機械的に抽出したものです。 再就職そのものは合法な届出に基づく公表データであり、 利益相反や違法行為を断定するものではありません。 追加検証には個別の届出内容、 在職時の所属部局、 再就職時の役職、 利害関係の有無等の確認が必要です。</p>")
lines.append("</body></html>")

html_path = OUTPUT_DIR / "digital_r7_jinjikyoku_cross_r5.html"
html_path.write_text("\n".join(lines), encoding="utf-8")
print("  HTML: {} ({:,} bytes)".format(html_path.relative_to(BASE), html_path.stat().st_size))

# サマリーテキスト (XポストやSESSION_LOG用)
summary_lines = []
summary_lines.append("# 内閣人事局 x デジタル庁R7 クロスリファレンス結果サマリー")
summary_lines.append("")
summary_lines.append("- データ: 内閣人事局 令和5年度公表分（{} 件）".format(len(all_records)))
summary_lines.append("- 期間: 令和5年4月1日 - 令和6年3月31日")
summary_lines.append("- 出典: https://www.cas.go.jp/jp/gaiyou/jimu/jinjikyoku/106-25-2/r06/kouhyou_0924.html")
summary_lines.append("- デジタル庁R7受注企業18社とのマッチ件数: **{} 件**".format(len(matches)))
summary_lines.append("")
summary_lines.append("## ベンダー別 集計")
summary_lines.append("")
if vendor_counts:
    summary_lines.append("| 受注企業 | 再就職件数 |")
    summary_lines.append("|---|---|")
    for vendor, count in sorted(vendor_counts.items(), key=lambda x: -x[1]):
        summary_lines.append("| {} | {} |".format(vendor, count))
else:
    summary_lines.append("該当なし（マッチング条件の見直しを検討）")
summary_lines.append("")
summary_lines.append("生成日時: " + __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

summary_path = OUTPUT_DIR / "digital_r7_jinjikyoku_cross_r5_SUMMARY.md"
summary_path.write_text("\n".join(summary_lines), encoding="utf-8")
print("  Summary: {} ({:,} bytes)".format(summary_path.relative_to(BASE), summary_path.stat().st_size))

print("")
print("=" * 70)
print("完了！")
print("=" * 70)
print("  open output/digital_r7_jinjikyoku_cross_r5.html  # ブラウザで確認")
print("  cat output/digital_r7_jinjikyoku_cross_r5_SUMMARY.md  # サマリー確認")
print("")
print("注意:")
print(" - マッチ0件の場合、 マッチング条件（株式会社/法人名フル表記）が厳しすぎる可能性")
print(" - R5公表分は『令和5年4月-令和6年3月の再就職』なので、 R7受注時点での在籍状況とは時間差あり")
print(" - 完全な照合には R7四半期データ（令和7年公表分）の追加が必要")
